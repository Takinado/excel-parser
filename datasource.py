import datetime
from pathlib import Path

from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet import worksheet


class DataSource:
    def read(self, file_name: Path) -> list[tuple]:
        workbook = load_workbook(file_name)
        sheet = workbook.worksheets[0]
        parsed_data: list = []
        for row in range(4, sheet.max_row + 1):
            for column in range(3, 11):
                value = sheet.cell(row, column).value
                date = sheet.cell(3, column).value
                data_type = self._get_data_type(sheet, column)
                company = sheet.cell(row, 2).value
                timing = self._get_timing(sheet, column)
                data_row = (
                    value,
                    self._get_date(date),
                    data_type,
                    company,
                    timing,
                )
                parsed_data.append(data_row)
        return parsed_data

    def _get_data_type(self, sheet: worksheet, column: int) -> str:
        return sheet.cell(row=2, column=column if column % 2 != 0 else column - 1).value

    def _get_timing(self, sheet: worksheet, column: int) -> str:
        return sheet.cell(row=1, column=3 if column in range(3, 7) else 7).value

    def _get_date(self, date: str) -> datetime.date:
        return (
            datetime.datetime.today().replace(day=1).date()
            if date == "data1"
            else datetime.datetime.today().replace(day=2).date()
        )
